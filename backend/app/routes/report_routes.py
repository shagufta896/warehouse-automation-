"""
Report export routes — download forecasts, inventory, and billing as Excel or PDF.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from app.database.database import get_db
from app.database.models import Product, SalesHistory, Bill, BillingItem, ForecastResult
from datetime import datetime, timedelta
import io, logging

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/reports", tags=["Reports"])


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")


# ── Inventory Stock Report ─────────────────────────────────────────────────
@router.get("/inventory/excel")
def inventory_excel(db: Session = Depends(get_db)):
    """Download full inventory as Excel workbook."""
    _require_openpyxl()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    products = db.query(Product).filter(Product.is_active == True).order_by(Product.category, Product.product_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1a1a2e")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Product ID", "Product Name", "Category", "Unit",
               "Current Stock", "Reorder Point", "Selling Price (₹)",
               "Cost Price (₹)", "Inventory Value (₹)", "Lead Time (days)", "Status"]
    col_widths = [12, 28, 18, 8, 14, 14, 18, 16, 20, 16, 14]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22

    for row_idx, p in enumerate(products, 2):
        stock  = p.current_stock or 0
        rp     = p.reorder_point
        status = "Out of Stock" if stock == 0 else ("Low Stock" if rp and stock <= rp else "OK")
        inv_val = round(stock * (p.selling_price or 0), 2)
        row_data = [
            p.product_id, p.product_name, p.category, p.unit or "pcs",
            stock, rp or "", p.selling_price or 0, p.cost_price or "",
            inv_val, p.supplier_lead_time or 5, status
        ]
        status_colors = {"OK": "d4edda", "Low Stock": "fff3cd", "Out of Stock": "f8d7da"}
        row_fill = PatternFill("solid", fgColor=status_colors.get(status, "FFFFFF"))
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx == 11:  # Status column
                cell.fill = row_fill
                cell.font = Font(bold=True)

    # Summary at bottom
    ws.append([])
    ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M"), "",
                "Total Products:", len(products), "",
                "Total Inventory Value:", f"₹{sum((p.current_stock or 0)*(p.selling_price or 0) for p in products):,.2f}"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"inventory_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Sales Report ──────────────────────────────────────────────────────────
@router.get("/sales/excel")
def sales_excel(days: int = 30, db: Session = Depends(get_db)):
    """Download sales history as Excel."""
    _require_openpyxl()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    since = datetime.utcnow() - timedelta(days=days)
    sales = db.query(SalesHistory).filter(SalesHistory.date >= since).order_by(desc(SalesHistory.date)).all()

    # Get product names
    product_map = {p.product_id: p.product_name for p in db.query(Product).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sales Last {days} Days"

    hdr_fill = PatternFill("solid", fgColor="0d6efd")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Date", "Product ID", "Product Name", "Units Sold", "Selling Price (₹)", "Revenue (₹)", "Day", "Season", "Festival", "Weekend", "Source"]
    widths   = [14,     12,            28,              12,            18,                 14,             12,    14,        14,          10,         10]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w

    total_revenue = 0
    for row_idx, s in enumerate(sales, 2):
        revenue = (s.units_sold or 0) * (s.selling_price or 0)
        total_revenue += revenue
        row_data = [
            s.date.strftime("%Y-%m-%d") if s.date else "",
            s.product_id,
            product_map.get(s.product_id, s.product_id),
            s.units_sold,
            s.selling_price,
            round(revenue, 2),
            s.day_of_week or "",
            s.season or "",
            s.festival or "",
            "Yes" if s.is_weekend else "No",
            s.source or "csv"
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    ws.append([])
    ws.append(["", "", "TOTAL", sum(s.units_sold or 0 for s in sales), "", round(total_revenue, 2)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"sales_{days}d_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Forecast Report ───────────────────────────────────────────────────────
@router.get("/forecast/excel")
def forecast_excel(product_id: str = None, db: Session = Depends(get_db)):
    """Download forecast results as Excel."""
    _require_openpyxl()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = db.query(ForecastResult)
    if product_id:
        q = q.filter(ForecastResult.product_id == product_id)
    forecasts = q.order_by(ForecastResult.product_id, ForecastResult.forecast_date).all()
    product_map = {p.product_id: p.product_name for p in db.query(Product).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecasts"

    hdr_fill = PatternFill("solid", fgColor="6f42c1")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["Product ID", "Product Name", "Forecast Date", "Predicted Demand", "Lower Bound", "Upper Bound", "Model Used"]
    widths   = [14,            28,              16,               16,                 14,             14,             14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_idx, f in enumerate(forecasts, 2):
        row_data = [
            f.product_id,
            product_map.get(f.product_id, f.product_id),
            f.forecast_date.strftime("%Y-%m-%d") if f.forecast_date else "",
            round(f.predicted_demand or 0, 1),
            round(f.lower_bound or 0, 1),
            round(f.upper_bound or 0, 1),
            f.model_used or ""
        ]
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"forecast_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Billing / Revenue Report ──────────────────────────────────────────────
@router.get("/billing/excel")
def billing_excel(days: int = 30, db: Session = Depends(get_db)):
    """Download billing/revenue report as Excel with summary sheet."""
    _require_openpyxl()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    since = datetime.utcnow() - timedelta(days=days)
    bills = db.query(Bill).filter(Bill.created_at >= since).order_by(desc(Bill.created_at)).all()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Bills ──
    ws1 = wb.active
    ws1.title = "Bills"
    hdr_fill = PatternFill("solid", fgColor="198754")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["Bill Number", "Date", "Customer", "Subtotal (₹)", "Discount (₹)", "GST (₹)", "Total (₹)", "Payment", "Status"]
    widths   = [18,            16,      22,           14,              14,             12,         12,          12,        12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws1.cell(row=1, column=i, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws1.column_dimensions[get_column_letter(i)].width = w

    for row_idx, b in enumerate(bills, 2):
        row_data = [
            b.bill_number,
            b.created_at.strftime("%Y-%m-%d %H:%M") if b.created_at else "",
            b.customer_name or "—",
            b.subtotal, b.discount, b.gst_amount, b.total,
            b.payment_method, b.status
        ]
        for col_idx, val in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=val).border = border

    # ── Sheet 2: Daily Revenue ──
    ws2 = wb.create_sheet("Daily Revenue")
    ws2.cell(1, 1, "Date").font = Font(bold=True)
    ws2.cell(1, 2, "Bills").font = Font(bold=True)
    ws2.cell(1, 3, "Revenue (₹)").font = Font(bold=True)
    daily = {}
    for b in bills:
        d = b.created_at.date() if b.created_at else None
        if d:
            daily.setdefault(d, {'bills': 0, 'revenue': 0})
            daily[d]['bills'] += 1
            daily[d]['revenue'] += b.total or 0
    for row_idx, (d, data) in enumerate(sorted(daily.items(), reverse=True), 2):
        ws2.cell(row_idx, 1, str(d))
        ws2.cell(row_idx, 2, data['bills'])
        ws2.cell(row_idx, 3, round(data['revenue'], 2))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"billing_{days}d_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Reorder Report ────────────────────────────────────────────────────────
@router.get("/reorder/excel")
def reorder_excel(db: Session = Depends(get_db)):
    """Download current reorder status for all products."""
    _require_openpyxl()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    products = db.query(Product).filter(Product.is_active == True).order_by(Product.category).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reorder Status"
    hdr_fill = PatternFill("solid", fgColor="dc3545")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["Product ID", "Product Name", "Category", "Current Stock", "Reorder Point", "Reorder Qty", "Lead Time (days)", "Status"]
    widths   = [14,            28,              18,           14,              14,              12,             16,               16]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_idx, p in enumerate(products, 2):
        stock = p.current_stock or 0
        rp    = p.reorder_point or 0
        status = "REORDER NOW" if (rp and stock <= rp) else ("OUT OF STOCK" if stock == 0 else "OK")
        fill_color = {"REORDER NOW": "fff3cd", "OUT OF STOCK": "f8d7da"}.get(status, "d4edda")
        row_data = [p.product_id, p.product_name, p.category, stock, rp or "—", p.reorder_quantity or "—", p.supplier_lead_time or 5, status]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx == 8:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"reorder_status_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
